"""Admin Router — model settings, data upload, audit logs"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from typing import Optional
import io, csv, uuid

from database import get_db
import models
from auth_utils import require_government

router = APIRouter()

# ── In-memory model settings ───────────────────────────────────────────────────
_model_settings = {
    "lookback_days": 30, "forecast_days": 7, "lstm_layers": 2, "lstm_units": 128,
    "learning_rate": 0.001, "dropout_rate": 0.2, "batch_size": 64, "epochs": 100,
    "num_features": 57, "num_basins": 70, "bias_correction": True,
}


@router.get("/model/settings")
def get_model_settings():
    return _model_settings


@router.post("/model/settings")
def save_model_settings(body: dict, _: models.User = Depends(require_government)):
    _model_settings.update(body)
    return {"message": "Settings saved", "settings": _model_settings}


# ── Data Upload ────────────────────────────────────────────────────────────────

def _safe_float(val):
    try:
        return float(val) if val not in (None, "", "N/A", "-") else None
    except:
        return None

def _safe_str(val):
    return str(val).strip() if val not in (None, "") else None


@router.post("/data/upload/{upload_type}")
async def upload_data(
    upload_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_government),
):
    contents = await file.read()
    fname = file.filename.lower()

    # ── Gauge Stations (CSV or XLSX) ──────────────────────────────────────────
    if upload_type == "gauges-csv":
        added, updated, skipped = 0, 0, 0
        errors = []

        try:
            # ── XLSX ──────────────────────────────────────────────────────────
            if fname.endswith(".xlsx") or fname.endswith(".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
                ws = wb.active

                # Find header row (look for S/N or NAME OF STATION)
                header_row = None
                for r in range(1, 15):
                    row_vals = [str(ws.cell(r, c).value or "").upper() for c in range(1, 10)]
                    if any("STATION" in v or "S/N" in v or "NAME" in v for v in row_vals):
                        # Confirm data starts next row
                        next = ws.cell(r+1, 1).value
                        if next is not None and str(next).strip() not in ("", "None"):
                            header_row = r
                            break

                if header_row is None:
                    raise ValueError("Could not find header row in the Excel file")

                # Map column positions from header
                headers = [str(ws.cell(header_row, c).value or "").upper().strip()
                           for c in range(1, ws.max_column + 1)]

                def col(keywords):
                    for i, h in enumerate(headers):
                        if any(k.upper() in h for k in keywords):
                            return i + 1
                    return None

                col_sn     = col(["S/N", "SN", "NO", "#"])
                col_river  = col(["RIVER"])
                col_name   = col(["STATION", "NAME OF STATION"])
                col_ha     = col(["HA", "CATEGORY", "CLASS"])
                col_lng    = col(["LONG", "LON", "LONGITUDE"])
                col_lat    = col(["LAT", "LATITUDE"])
                col_state  = col(["STATE"])
                col_warn   = col(["WARN", "WARNING"])
                col_danger = col(["DANGER", "CRITICAL"])

                for r in range(header_row + 1, ws.max_row + 1):
                    sn       = ws.cell(r, col_sn).value   if col_sn    else None
                    river    = ws.cell(r, col_river).value if col_river else None
                    name     = ws.cell(r, col_name).value  if col_name  else None
                    ha       = ws.cell(r, col_ha).value    if col_ha    else None
                    lng_val  = ws.cell(r, col_lng).value   if col_lng   else None
                    lat_val  = ws.cell(r, col_lat).value   if col_lat   else None
                    state    = ws.cell(r, col_state).value if col_state else None
                    warn     = ws.cell(r, col_warn).value  if col_warn  else None
                    danger   = ws.cell(r, col_danger).value if col_danger else None

                    if not name or str(name).strip() in ("", "None", "NAME OF STATION"):
                        continue

                    name    = _safe_str(name)
                    river   = _safe_str(river)
                    lat     = _safe_float(lat_val)
                    lng     = _safe_float(lng_val)
                    station_code = f"S{int(sn):04d}" if sn else "S" + str(uuid.uuid4())[:6].upper()

                    # Validate coords are within Nigeria
                    if lat is not None and lng is not None:
                        if not (4.0 <= lat <= 14.0) or not (2.5 <= lng <= 15.0):
                            # Try swapping lat/lng (common data entry mistake)
                            if (4.0 <= lng <= 14.0) and (2.5 <= lat <= 15.0):
                                lat, lng = lng, lat
                            else:
                                errors.append(f"Row {r}: {name} — coords out of range ({lat}, {lng})")
                                skipped += 1
                                continue

                    existing = db.query(models.RiverGauge).filter(
                        models.RiverGauge.station_name == name,
                        models.RiverGauge.river_name == river
                    ).first()

                    if existing:
                        # Update coords if missing
                        if lat and not existing.lat: existing.lat = lat
                        if lng and not existing.lng: existing.lng = lng
                        if state and not existing.state: existing.state = _safe_str(state)
                        updated += 1
                    else:
                        db.add(models.RiverGauge(
                            station_code=station_code,
                            station_name=name,
                            river_name=river,
                            lat=lat,
                            lng=lng,
                            state=_safe_str(state),
                            is_active=True,
                            warning_level=_safe_float(warn),
                            danger_level=_safe_float(danger),
                            metadata_json={"ha_category": _safe_str(ha)} if ha else None,
                        ))
                        added += 1

                db.commit()

            # ── CSV ───────────────────────────────────────────────────────────
            elif fname.endswith(".csv"):
                text = contents.decode("utf-8-sig", errors="replace")
                reader = csv.DictReader(io.StringIO(text))
                for i, row in enumerate(reader):
                    # Normalise keys
                    row = {k.strip().upper(): v for k, v in row.items()}

                    name  = (row.get("NAME OF STATION") or row.get("STATION NAME") or row.get("STATION") or "").strip()
                    river = (row.get("NAME OF RIVER") or row.get("RIVER NAME") or row.get("RIVER") or "").strip()
                    lat   = _safe_float(row.get("LAT.") or row.get("LAT") or row.get("LATITUDE"))
                    lng   = _safe_float(row.get("LONG.") or row.get("LONG") or row.get("LON") or row.get("LONGITUDE"))
                    state = _safe_str(row.get("STATE"))

                    if not name:
                        skipped += 1
                        continue

                    if lat and lng:
                        if not (4.0 <= lat <= 14.0) or not (2.5 <= lng <= 15.0):
                            if (4.0 <= lng <= 14.0) and (2.5 <= lat <= 15.0):
                                lat, lng = lng, lat

                    existing = db.query(models.RiverGauge).filter(
                        models.RiverGauge.station_name == name
                    ).first()

                    if existing:
                        if lat and not existing.lat: existing.lat = lat
                        if lng and not existing.lng: existing.lng = lng
                        updated += 1
                    else:
                        code = "S" + str(uuid.uuid4())[:6].upper()
                        db.add(models.RiverGauge(
                            station_code=code, station_name=name, river_name=river or None,
                            lat=lat, lng=lng, state=state, is_active=True,
                            warning_level=_safe_float(row.get("WARNING LEVEL") or row.get("WARN")),
                            danger_level=_safe_float(row.get("DANGER LEVEL") or row.get("DANGER")),
                        ))
                        added += 1

                db.commit()
            else:
                return {"message": f"Unsupported file type. Please upload .xlsx or .csv", "success": False}

        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

        msg_parts = []
        if added:   msg_parts.append(f"{added} stations added")
        if updated: msg_parts.append(f"{updated} existing stations updated")
        if skipped: msg_parts.append(f"{skipped} rows skipped")
        if errors:  msg_parts.append(f"{len(errors)} coordinate errors")

        return {
            "message": ", ".join(msg_parts) + "." if msg_parts else "No new stations found.",
            "added": added, "updated": updated, "skipped": skipped, "errors": errors[:10],
        }

    # ── Gauge Readings CSV ────────────────────────────────────────────────────
    elif upload_type == "readings-csv":
        from datetime import datetime
        added, skipped = 0, 0
        try:
            text = contents.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                row = {k.strip().upper(): v for k, v in row.items()}
                sid   = (row.get("STATION ID") or row.get("STATION_ID") or "").strip()
                level = _safe_float(row.get("WATER LEVEL") or row.get("LEVEL"))
                if not sid or level is None:
                    skipped += 1; continue
                gauge = db.query(models.RiverGauge).filter(
                    (models.RiverGauge.station_code == sid) | (models.RiverGauge.station_id == sid)
                ).first()
                if not gauge:
                    skipped += 1; continue
                db.add(models.GaugeReading(
                    station_id=gauge.station_id, water_level=level,
                    discharge=_safe_float(row.get("FLOW RATE") or row.get("DISCHARGE")),
                    recorded_at=datetime.utcnow(), source="upload",
                ))
                added += 1
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=str(e))
        return {"message": f"{added} readings added, {skipped} skipped.", "added": added}

    # ── Other upload types (shapefiles etc) — queue for later ─────────────────
    else:
        size_kb = len(contents) / 1024
        return {
            "message": f"File '{file.filename}' received ({size_kb:.1f} KB). Shapefile processing will be available in the next update.",
            "upload_type": upload_type,
        }


# ── Audit Logs ─────────────────────────────────────────────────────────────────
@router.get("/audit/logs")
def get_audit_logs(db: Session = Depends(get_db), limit: int = 100, _: models.User = Depends(require_government)):
    logs = db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc()).limit(limit).all()
    return [
        {
            "id": l.id,
            "user_email": l.user_id,
            "action": l.action,
            "resource_type": l.resource,
            "resource_id": l.resource_id,
            "details": str(l.details) if l.details else "",
            "ip_address": l.ip_address,
            "created_at": l.created_at.isoformat() if l.created_at else None,
        }
        for l in logs
    ]
